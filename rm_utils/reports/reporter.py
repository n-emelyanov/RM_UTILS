from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

import pandas as pd
import numpy as np
import os


class ExcelReporter:
    """
    Класс для формирования отчетов в Excel в стилистике GPB.


    Parameters
    ----------
    file_name : str
        Путь к фалу Excel.

    Examples
    --------
    >>>> data = pd.read_csv('data.csv')
    >>>> writer = ExcelReport('report.xlsx')
    >>>> writer.add_dataframe(data, sheet_name='Sheet', index=True, row_offset=5, col_offset=2)
    >>>> writer.add_image('image.jpg', sheet_name='Sheet', row_offset=5, col_offset=7, scale=0.7)
    >>>> writer.save()

    """

    def __init__(self, file_name: str):
        self.file_name = file_name
        self.workbook = Workbook()
        print(self.workbook.sheetnames)
        self.workbook.remove_sheet(self.workbook.get_sheet_by_name('Sheet'))

        # self._blue_fill = PatternFill(fill_type='solid', fgColor='2353D6') # GPB
        self._blue_fill = PatternFill(fill_type='solid', fgColor='7000FF') # UZUM

        self._dark_blue_fill = PatternFill(fill_type='solid', fgColor='113D70')

        self._grid_border = Border(left=Side(style='thin', color='D9D9D9'),
                                   right=Side(style='thin', color='D9D9D9'),
                                   top=Side(style='thin', color='D9D9D9'),
                                   bottom=Side(style='thin', color='D9D9D9'))

        self._thin_border = Border(left=Side(style='thin'),
                                   right=Side(style='thin'),
                                   top=Side(style='thin'),
                                   bottom=Side(style='thin'))



    def _prepare_sheet(self, sheet_name: str = 'Sheet') -> Worksheet:
        """
        Изменяет стиль листа (GPB_style).
        """

        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
            sheet = self.workbook[sheet_name]

            # Расширяем первые 2 строки и заливаем синим
            for i in range(1, 3):
                sheet.row_dimensions[i].height = 22
                sheet.row_dimensions[i].fill = self._blue_fill

            # Добавляем лого GPB
            image_path = os.path.join(os.path.dirname(__file__), '..', 'assets', 'logo.png')
            # self.add_image(image_path, sheet_name, row_offset=1.2, col_offset=1.059354, scale=0.7) # GPB
            self.add_image(image_path, sheet_name, row_offset=1.2, col_offset=1.059354, scale=0.3) # UZUM

            # Убираем стандартную сетку excel
            sheet.sheet_view.showGridLines = False

        return self.workbook[sheet_name]



    def _merge_multiindex_cells(self, data: pd.DataFrame, sheet: Worksheet, row_offset: int, col_offset: int) -> None:
        """
        Соединяет ячейки для первого столбца мультииндекса.
        """
        index_levels = data.index.levels[0]

        for level in index_levels:
            start_row = data.index.get_loc_level(level, level=0)[0].start + row_offset + 1
            end_row = start_row + len(data.loc[level]) - 1

            sheet.merge_cells(start_row=start_row,
                              start_column=col_offset,
                              end_row=end_row,
                              end_column=col_offset)

            # Получаем замердженную ячейку
            cell = sheet.cell(row=start_row, column=col_offset)

            # Выравниваем по центру
            cell.alignment = Alignment(horizontal='center', vertical='center')



    def _add_borders(self, sheet: Worksheet, last_row: int, last_column: int, row_offset: int, col_offset: int) -> None:
        """
        Добавляет внешние границы вокруг датафрейма.
        """

        thin = Side(border_style="thin", color="000000")

        # Функция для установки границ
        def set_border(row: int, col: int, direction: str) -> None:
            original_border = {key: value for key, value in sheet.cell(row=row, column=col).border.__dict__.items() if key != direction}
            if direction == 'top':
                sheet.cell(row=row, column=col).border = Border(top=thin, **original_border)
            elif direction == 'bottom':
                sheet.cell(row=row, column=col).border = Border(bottom=thin, **original_border)
            elif direction == 'left':
                sheet.cell(row=row, column=col).border = Border(left=thin, **original_border)
            else: # direction == 'right'
                sheet.cell(row=row, column=col).border = Border(right=thin, **original_border)

        # Установка границ для верхней и нижней строк
        for col in range(col_offset, last_column + 1):
            set_border(row_offset + 1, col, 'top')
            set_border(last_row, col, 'bottom')

        # Установка границ для левого и правого столбцов
        for row in range(row_offset + 1, last_row + 1):
            set_border(row, col_offset, 'left')
            set_border(row, last_column, 'right')



    def add_dataframe(
        self,
        data: pd.DataFrame,
        sheet_name: str = 'Sheet',
        table_name: str = None,
        index: bool = False,
        row_offset: int = 5,
        col_offset: int = 2
    ) -> None:
        """
        Добавляет датафрейм в файл Excel.

        Parameters
        ----------
        data : pd.DataFrame
            Датафрейм для вставки.

        sheet_name : str, default='Sheet'
            Название листа. Если указанного листа не существует - создает новый лист.

        table_name : str, default=None
            Название таблицы. Выводится заглавным шрифтом над вставляемым датафреймом.

        index : bool, default=False
            Отвечает за вставку индекса датафрейма.

        row_offset : int, default=1
            Отступ по строкам.

        col_offset : int, default=1
            Отступ по столбцам.

        """

        # Форматируем лист (GPB style)
        sheet = self._prepare_sheet(sheet_name)

        # Если указан параметр table_name - добавляем название таблицы
        if table_name:
            cell = sheet.cell(row=row_offset, column=col_offset, value=table_name)
            cell.font = Font(size=14, bold=True)
            row_offset += 1

        # Добавляем датафрейм
        for row_id, row in enumerate(dataframe_to_rows(data, index=index, header=True)):
            for col_id, cell_value in enumerate(row):

            # Если индекс - пропускаем вторую (пустую) строку
                if index and row_id == 1:
                    row_offset -= 1
                    break

                cell = sheet.cell(row=row_offset + row_id, column=col_offset + col_id)

                if isinstance(cell_value, (list, np.ndarray, pd.Period, np.datetime64, pd.Interval)):
                    cell.value = str(cell_value)
                else:
                    cell.value = cell_value

                # Проставляем жирный шрифт для столбцов
                if row_id == 0 and cell.value:
                    cell.font = Font(bold=True, color='00FFFFFF')
                    cell.fill = self._dark_blue_fill
                    cell.border = self._thin_border

                # Проставляем границы для индекса (первый столбец)
                elif index and col_id == 0 and row_id != 0:
                    cell.border = self._thin_border
                    cell.font = Font(bold=True)

                # Заполняем серыми границами (а-ля grid_lines) ячейки со значениями
                elif row_id != 0:
                    cell.border = self._grid_border

        # Костыль для корректной вставки внешних границ
        if index:
            row_offset += 1
            row_id -= 1

        # Мерджим ячейки с мультииндексом
        if index and isinstance(data.index, pd.core.indexes.multi.MultiIndex):
            self._merge_multiindex_cells(data, sheet, row_offset, col_offset)

        # Добавляем внешние границы вокруг датафрейма
        self._add_borders(sheet,
                          last_row=row_offset + row_id,
                          last_column=col_offset + col_id,
                          row_offset=row_offset,
                          col_offset=col_offset)



    def add_image(
        self,
        img_path: str,
        sheet_name: str = 'Sheet',
        row_offset: float = 1,
        col_offset: float = 1,
        scale: float = 1
    ) -> None:
        """
        Добавляет изображение в файл Excel.

        Parameters
        ----------
        image_path : str
            Пусть к изображению.

        sheet_name : str, default='Sheet'
            Название листа. Если указанного листа не существует - создает новый лист.

        row_offset : float, default=1
            Отступ по строкам.

        col_offset : float, default=1
            Отступ по столбцам.

        scale: float, default=1
            Определяет размер вставляемого изображения.

        Notes
        -----
        В данной функции отступы (row_offset, col_offset) можно подавать в формате float
        на случай, если вы например захотите вставить изображение начиная с середины ячейки.

        """
        # Форматируем лист (GPB style)
        sheet = self._prepare_sheet(sheet_name)

        img = Image(img_path)

        h = img.height * scale
        w = img.width * scale

        size = XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h))

        cellh = lambda x: cm_to_EMU((x * 49.77)/99)
        cellw = lambda x: cm_to_EMU((x * (18.65-1.71))/10)

        col_bias, row_bias = cellw(col_offset%1), cellh(row_offset%1)
        col_offset, row_offset = int(col_offset), int(row_offset)

        marker = AnchorMarker(col=col_offset - 1,
                              colOff=col_bias,
                              row=row_offset - 1,
                              rowOff=row_bias)

        img.anchor = OneCellAnchor(_from=marker, ext=size)

        sheet.add_image(img)



    def save(self):
        """
        Сохраняет итоговый Excel файл.
        """
        self.workbook.save(self.file_name)
